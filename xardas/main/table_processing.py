import io
import openpyxl
from collections import namedtuple
from django.conf import settings
from django.core.exceptions import FieldDoesNotExist
from openpyxl.drawing.image import Image
from .xls_map_character import CHARACTER_FORM_RECORDS, IMAGE_SIZES
from .utilites import get_date_time
from .models.character import CHARACTER_NAME_FIELD


class BaseKeyNotFound(KeyError):
    pass


class ExportXLS:
    DOCUMENT_EXTENSION = '.xlsx'
    CONTENT_TYPE = 'application/doc'

    def __init__(self, char):
        self.char = char
        self.doc_name = self.get_doc_name()
        self.id_file = None
        self.ws = None

    def __enter__(self):
        self.id_file = io.BytesIO()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.id_file:
            self.id_file.close()

    def xls_insert_data(self, context):
        for key, value in context.items():
            try:
                self.ws[key] = value
            except AttributeError:
                print(f'TEMPLATE ERROR: Key: {key} Value: {value}')

    def xls_insert_image(self, xls_cell, path_to_image, db_field):
        print(f'Xls cell "{xls_cell}" value: "{path_to_image}"')
        image = openpyxl.drawing.image.Image(path_to_image)

        try:
            print(
                f'Номер колонки: "{self.ws[xls_cell].column}"\r\nНомер строки: "{self.ws[xls_cell].row}"\r\n Буква строки:'
                f' "{self.ws[xls_cell].column_letter}"\r\nШирина изображения: "{image.width}"\r\nВысота изображения: "{image.height}"\r\n'
                f'Путь к изображению: {path_to_image}')

            row = self.ws[xls_cell].row
            column = self.ws[xls_cell].column_letter

            print(f'\r\nColumn column_dimensions: {self.ws.column_dimensions[column].width} \r\n')
            print(f'Column row_dimensions: {self.ws.row_dimensions[row].height} \r\n')

        except TypeError as error:
            print(f'Error xls_insert_image: {error}')

        for field_name, image_size in IMAGE_SIZES.items():
            if field_name == db_field:
                print(f'Found field name: {db_field} width: {image_size[0]}, height: {image_size[1]}')
                image.width = image_size[0]
                image.height = image_size[1]
                break
        self.ws.add_image(image, xls_cell)

    def generate_xls(self):
        workbook = openpyxl.load_workbook(settings.XLS_TEMPLATE_PATH)
        sheet_index = 0
        work_book_sheet_names = workbook.sheetnames

        for template_records in CHARACTER_FORM_RECORDS:
            print(f'Страница: {work_book_sheet_names[sheet_index]}\r\nЗаписи шаблона: {template_records}\r\n')
            # openpyxl.utils.exceptions.InvalidFileException:
            self.ws = workbook[work_book_sheet_names[sheet_index]]
            context = self.make_form_data(template_records)
            self.xls_insert_data(context)

            sheet_index+=1
        workbook.save(self.id_file)
        self.id_file.seek(0)
        return self.doc_name, self.id_file

    def make_form_data(self, form_records) -> dict:
        data = {}
        DOC_FORM = namedtuple('DOC_RECORDS', 'db_field xls_cell')
        for _record in form_records:
            record = DOC_FORM(*_record)
            if record.xls_cell:
                print(f'DB_field: "{record.db_field}" -> XLS_field: "{record.xls_cell}"')
                value = self.get_db_value(record.db_field, record.xls_cell)
                if value:
                    data[record.xls_cell] = value
        return data

    def get_doc_name(self):
        character_name = self.get_db_value(CHARACTER_NAME_FIELD)
        if not character_name:
            raise BaseKeyNotFound
        part_name = get_date_time('%Y%m%d')
        return character_name+part_name + self.DOCUMENT_EXTENSION

    def get_db_value(self, db_field, xls_cell = None):
        value = ''
        field_type = 'Unknown'
        field_verbose = 'Unknown'

        try:
            field_type = self.get_type_field(db_field)
            field_verbose = self.get_verbose_field(db_field)
        except FieldDoesNotExist as error:
            print(f'Field type error: {db_field} - {error}')

        try:
            value = getattr(self.char, db_field)
        except AttributeError as error:
            print(f'get_db_value error: {error}')

        print(f'Value: "{value}" Description: "{field_verbose}" Type: "{field_type}"\r\n')

        if field_type == 'BooleanField' and value == True:
            value = '\u2714'  # ✔

        if field_type == 'ManyToManyField':
            if db_field == 'races':
                value = self.char.get_race()

        if field_type == 'FileField' and value:
            self.xls_insert_image(xls_cell, value, db_field)
            value = ''

        if value:
            try:
                value = str(value)
            except TypeError as err:
                print('Bad value: {}'.format(err))
                return False
            return value

    def get_type_field(self, db_field) -> str:
        result = self.char._meta.get_field(db_field).get_internal_type()
        return result

    def get_verbose_field(self, db_field) -> str:
        result = self.char._meta.get_field(db_field).verbose_name.title()
        return result
